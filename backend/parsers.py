import xylib
from util import KnownPeak
import re
import numpy as np
from openpyxl import Workbook
class SpectrumParser:
    """Parser for all Spectrum file formats"""
    def __init__(self, fname):
        self.fname = fname
        if self.fname.split(".")[-1].lower() == "spe":
            self.speFile = True
        else:
            self.speFile = False
    def getValues(self):
        if self.speFile:
            self.spectrumfile = open(self.fname)
            text = self.spectrumfile.read()
            self.spectrumfile.close()
            l = text.split('$')[1:]
            sections = dict()
            for section in l:
                sp = section.split('\n',1)
                sections[sp[0]] = sp[1]
            livetime, realtime = sections["MEAS_TIM:"].strip().split(" ")
            livetime = float(livetime)
            realtime = float(realtime)
            startind, endind = sections["DATA:"].split("\n")[0].split(" ")
            data = sections["DATA:"].split("\n")[1:-1]
            data = [int(i) for i in data]
            intercept, slope = sections["ENER_FIT:"].strip().split(" ")[:2]
            intercept = float(intercept)
            slope = float(slope)
            energies = [intercept + i*slope for i in range(int(startind),int(endind)+1)]
            cps = [total/livetime for total in data]
            return {"livetime":livetime, "realtime":realtime, "energies":np.array(energies), "cps":np.array(cps)}
        else:
            d = xylib.load_file(self.fname, '')
            block = d.get_block(0)
            meta = block.meta
            metaDict = {}
            for i in range(meta.size()):
                key = meta.get_key(i)
                metaDict[key] = meta.get(key)
            ncol = block.get_column_count()
            nrow = block.get_point_count()
            data = [[block.get_column(i).get_value(j) for j in range(nrow)] for i in range(1, ncol+1)]
            livetime = metaDict["live time (s)"]
            realtime = metaDict["real time (s)"]
            return {"livetime":livetime, "realtime":realtime, "energies":np.array(data[0]), "cps":np.array(data[1])/float(livetime)}

class StandardsFileParser:
    def __init__(self,fname=""):
        self.file = open(fname, "r")
        self.peaks = None
    def extract_peaks(self, delayed):
        if self.peaks != None:
            return self.peaks
        lines = self.file.readlines()
        headings = re.sub(r'[^\x00-\x7F]+','', lines[0]).strip().split(",")
        lines = [l.split(",") for l in lines[1:]]
        try:
            peak_energy_index = headings.index("Energy (keV)")
            isotope_index = headings.index("Isotope")
        except:
            raise ValueError("Bad Sensitivity File Format.")
        re_mass = re.compile(r"Mass \((\w+)\)")
        re_sens = re.compile (r"Sensitivity \(cps/(\w+)\)")
        reHalfLife = re.compile(r"[Hh]alf-[Ll]ife \((\w+)\)")
        reDecayConstant = re.compile(r"[Dd]ecay [Cc]onstant \(1/(\w+)\)")
        halfLifeIndex=None
        decayConstantIndex = None
        decayConstant = None
        decayUnit = None

        sensitivity_divisor = None
        divisor_index = None
        unit = None
        for i,h in enumerate(headings):
            if re_mass.match(h) != None:
                sensitivity_divisor = False
                divisor_index = i
                unit = re_mass.match(h).group(1)
            elif re_sens.match(h) != None:
                sensitivity_divisor = True
                divisor_index = i
                unit = re_sens.match(h).group(1)
            elif reDecayConstant.match(h):
                decayConstant = True
                decayUnit = reDecayConstant.match(h).group(1)
                decayConstantIndex = i
            elif reHalfLife.match(h):
                decayConstant = False
                decayUnit = reHalfLife.match(h).group(1)
                halfLifeIndex = i
        if delayed:
            if decayConstant == None:
                raise ValueError("Delayed analysis, must provide half-life or decay constant in sensitivity file")
            elif decayConstant:
                if sensitivity_divisor == None:
                    self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index]), decayConstant=l[decayConstantIndex], decayUnit = decayUnit) for l in lines]
                elif sensitivity_divisor:
                    self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index]), sensitivity = float(l[divisor_index]), unit=unit, decayConstant=l[decayConstantIndex], decayUnit = decayUnit) for l in lines]
                else:
                    self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index]), mass = float(l[divisor_index]), unit=unit, decayConstant=l[decayConstantIndex], decayUnit = decayUnit) for l in lines]
            else:
                if sensitivity_divisor == None:
                    self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index]), halfLife=l[halfLifeIndex], decayUnit = decayUnit) for l in lines]
                elif sensitivity_divisor:
                    self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index]), sensitivity = float(l[divisor_index]), unit=unit, halfLife=l[halfLifeIndex], decayUnit = decayUnit) for l in lines]
                else:
                    self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index]), mass = float(l[divisor_index]), unit=unit, halfLife=l[halfLifeIndex], decayUnit = decayUnit) for l in lines]
        else:
            if sensitivity_divisor == None:
                self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index])) for l in lines]
            elif sensitivity_divisor:
                self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index]), sensitivity = float(l[divisor_index]), unit=unit) for l in lines]
            else:
                self.peaks = [KnownPeak(l[isotope_index],float(l[peak_energy_index]), mass = float(l[divisor_index]), unit=unit) for l in lines]
        self.file.close()
        return self.peaks

class CSVWriter:
    def __init__(self, projectID, fname, headings, data):
        self.fname = "./results/" + projectID + "/" + fname
        self.headings = headings
        self.data = data
    def write(self):
        f = open(self.fname, "w")
        f.seek(0)
        f.write(",".join(self.headings)+"\n")
        for line in self.data:
            try:
                ld = [str(e) for e in line[0]]
            except:
                ld = [str(e) for e in line]
            f.write(','.join(ld)+"\n")
        f.close()

class ExcelWriter:
    def __init__(self, projectID, projectTitle, allFilenames, headings, data):
        self.fname = "./results/" + projectID + "/" + projectTitle.replace(" ","_").replace("\\","") + ".xlsx"
        self.allFilenames = allFilenames
        self.headings = headings
        self.data = data
    def write(self):
        wb = Workbook()

        ws = wb.active
        ws.title = "All Files"
        ws["A1"] = "Filename"
        for i in range(len(self.headings[0][0])):
            _ = ws.cell(row=1, column=i+2, value=self.headings[0][0][i])
        rowCount = 1
        for i in range(len(self.allFilenames)):
            for j in range(len(self.data[i][0])):
                _ = ws.cell(row=rowCount+1, column=1, value=self.allFilenames[i].split("\\")[-1])
                for k in range(len(self.data[i][0][j])):
                    _ = ws.cell(row=rowCount+1, column=k+2, value=self.data[i][0][j][k])
                rowCount += 1

        for i in range(len(self.allFilenames)):
            newWs = wb.create_sheet(self.allFilenames[i].split("\\")[-1][:31])
            for l in range(len(self.headings[0][0])):
                _ = newWs.cell(row=1, column=l+1, value=self.headings[0][0][l])
            for j in range(len(self.data[i][0])):
                for k in range(len(self.data[i][0][j])):
                    _ = newWs.cell(row=j+2, column=k+1, value=self.data[i][0][j][k])
        
        wb.save(self.fname)



